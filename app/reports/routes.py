import io
from flask import render_template, redirect, url_for, flash, session, request, send_file
from . import bp
from ..models import db, Item, Sale
from ..services import get_profit_for_period, get_best_sellers, get_worst_sellers, get_attendant_performance, get_attendant_sales_details, get_sales_trend_data
from ..ai_service import generate_sales_insights_narrative
from datetime import datetime, timedelta

@bp.before_request
def require_owner():
    if 'staff_id' not in session:
        return redirect(url_for('auth.login'))
    if session.get('role') != 'owner':
        flash('Access denied. Owners only.', 'danger')
        return redirect(url_for('sales.dashboard'))

def get_date_range(period):
    now = datetime.utcnow()
    if period == 'today':
        return now.replace(hour=0, minute=0, second=0, microsecond=0), None
    elif period == 'week':
        start = now - timedelta(days=now.weekday())
        return start.replace(hour=0, minute=0, second=0, microsecond=0), None
    return None, None

@bp.route('/')
def index():
    shop_id = session.get('shop_id')
    period = request.args.get('period', 'all')
    start_date, end_date = get_date_range(period)
    
    profit = get_profit_for_period(shop_id, start_date, end_date)
    best_sellers = get_best_sellers(shop_id, start_date=start_date, end_date=end_date)
    worst_sellers = get_worst_sellers(shop_id, start_date=start_date, end_date=end_date)
    
    return render_template('reports/index.html', profit=profit, best_sellers=best_sellers, worst_sellers=worst_sellers, period=period)

@bp.route('/attendants')
def attendants():
    shop_id = session.get('shop_id')
    period = request.args.get('period', 'all')
    start_date, end_date = get_date_range(period)
    
    performance = get_attendant_performance(shop_id, start_date, end_date)
    details = get_attendant_sales_details(shop_id, start_date, end_date)
    
    return render_template('reports/attendants.html', performance=performance, details=details, period=period)

@bp.route('/insights')
def insights():
    shop_id = session.get('shop_id')
    trend_data = get_sales_trend_data(shop_id)
    lang = session.get('lang', 'en')
    narrative = generate_sales_insights_narrative(trend_data, lang=lang)
    
    return render_template('reports/insights.html', trend_data=trend_data, narrative=narrative)

@bp.route('/export')
def export_sales():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    shop_id = session.get('shop_id')
    period = request.args.get('period', 'all')
    start_date, end_date = get_date_range(period)
    
    query = Sale.query.join(Item).filter(Item.shop_id == shop_id)
    if start_date:
        query = query.filter(Sale.timestamp >= start_date)
    if end_date:
        query = query.filter(Sale.timestamp <= end_date)
        
    sales = query.order_by(Sale.timestamp.desc()).all()
    
    wb = Workbook()
    
    # 1. All Sales
    ws_all = wb.active
    ws_all.title = "All Sales"
    headers = ["Date", "Item", "Category", "Quantity", "Price", "Total", "Attendant", "Client"]
    ws_all.append(headers)
    
    best_sale = None
    worst_sale = None
    
    for sale in sales:
        client_name = sale.client.name if sale.client else "Walk-in"
        attendant_name = sale.staff.name if sale.staff else "Unknown"
        
        row = [
            sale.timestamp.strftime('%Y-%m-%d %H:%M'),
            sale.item.name,
            sale.item.category or "",
            sale.quantity_sold,
            float(sale.price_at_sale),
            float(sale.total_amount),
            attendant_name,
            client_name
        ]
        ws_all.append(row)
        
        if not best_sale or sale.total_amount > best_sale.total_amount:
            best_sale = sale
        if not worst_sale or sale.total_amount < worst_sale.total_amount:
            worst_sale = sale
            
    for cell in ws_all["1:1"]:
        cell.font = Font(bold=True)
        
    # 2. Best Sale
    ws_best = wb.create_sheet("Best Sale")
    ws_best.append(headers)
    if best_sale:
        ws_best.append([
            best_sale.timestamp.strftime('%Y-%m-%d %H:%M'),
            best_sale.item.name,
            best_sale.item.category or "",
            best_sale.quantity_sold,
            float(best_sale.price_at_sale),
            float(best_sale.total_amount),
            best_sale.staff.name if best_sale.staff else "Unknown",
            best_sale.client.name if best_sale.client else "Walk-in"
        ])
    for cell in ws_best["1:1"]:
        cell.font = Font(bold=True)
        
    # 3. Worst Sale
    ws_worst = wb.create_sheet("Worst Sale")
    ws_worst.append(headers)
    if worst_sale:
        ws_worst.append([
            worst_sale.timestamp.strftime('%Y-%m-%d %H:%M'),
            worst_sale.item.name,
            worst_sale.item.category or "",
            worst_sale.quantity_sold,
            float(worst_sale.price_at_sale),
            float(worst_sale.total_amount),
            worst_sale.staff.name if worst_sale.staff else "Unknown",
            worst_sale.client.name if worst_sale.client else "Walk-in"
        ])
    for cell in ws_worst["1:1"]:
        cell.font = Font(bold=True)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, as_attachment=True, download_name=f"sales_export_{period}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
